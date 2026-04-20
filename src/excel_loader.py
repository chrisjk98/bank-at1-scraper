from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from .models import BankRecord


@dataclass
class ExcelLoader:
    cache: dict[str, BankRecord] | None = None

    def __post_init__(self) -> None:
        if self.cache is None:
            self.cache = {}

    def load_banks(self, excel_path: str | Path, limit: int | None = None) -> list[BankRecord]:
        path = Path(excel_path)
        if not path.exists():
            raise FileNotFoundError(f"Excel file not found: {path}")
        if path.suffix.lower() not in {".xlsx", ".xls"}:
            raise ValueError("Input file must be .xlsx or .xls")

        rows = self._read_rows(path)
        records: list[BankRecord] = []
        for row in rows:
            gvkey = str(row.get("gvkey", "")).strip()
            company_name = str(row.get("company_name", "")).strip()
            if not gvkey or not company_name:
                continue
            record = BankRecord(gvkey=gvkey, company_name=company_name)
            self.cache[gvkey] = record
            records.append(record)
            if limit and len(records) >= limit:
                break
        return records

    def _read_rows(self, path: Path) -> Iterable[dict[str, str]]:
        try:
            import pandas as pd
        except ImportError:
            return self._read_rows_without_pandas(path)

        read_kwargs = {}
        if path.suffix.lower() == ".xls":
            read_kwargs["engine"] = "xlrd"

        try:
            df = pd.read_excel(path, **read_kwargs)
        except Exception:
            # fallback when headers are not present
            df = pd.read_excel(path, header=None, **read_kwargs)
            if len(df.columns) < 2:
                raise ValueError("Excel file must contain at least two columns: GVKey and Company Name")
            df = df.iloc[:, :2]
            df.columns = ["gvkey", "company_name"]

        normalized_columns: dict[str, str] = {}
        for column in df.columns:
            key = str(column).strip().lower().replace(" ", "")
            if key in {"gvkey", "gv_key", "a"}:
                normalized_columns[column] = "gvkey"
            elif key in {"conm", "companyname", "bankname", "b"}:
                normalized_columns[column] = "company_name"

        if "gvkey" not in normalized_columns.values() or "company_name" not in normalized_columns.values():
            first_two = list(df.columns[:2])
            if len(first_two) < 2:
                raise ValueError("Could not identify GVKey and Company Name columns")
            df = df[first_two]
            df.columns = ["gvkey", "company_name"]
        else:
            df = df.rename(columns=normalized_columns)
            df = df[["gvkey", "company_name"]]

        df = df.fillna("")
        return df.to_dict(orient="records")

    def _read_rows_without_pandas(self, path: Path) -> Iterable[dict[str, str]]:
        if path.suffix.lower() == ".xlsx":
            try:
                from openpyxl import load_workbook
            except ImportError as exc:  # pragma: no cover
                raise RuntimeError("Install openpyxl (or pandas) to read .xlsx files.") from exc

            workbook = load_workbook(path, read_only=True, data_only=True)
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                return []

            header = [str(value).strip().lower().replace(" ", "") if value is not None else "" for value in rows[0]]
            use_header = "gvkey" in header or "conm" in header or "companyname" in header
            data_rows = rows[1:] if use_header else rows

            results: list[dict[str, str]] = []
            for row in data_rows:
                if not row:
                    continue
                gvkey = row[0] if len(row) > 0 else ""
                company_name = row[1] if len(row) > 1 else ""
                results.append({"gvkey": str(gvkey or ""), "company_name": str(company_name or "")})
            return results

        if path.suffix.lower() == ".xls":
            raise RuntimeError("Install pandas with xlrd support to read .xls files.")
        return []
